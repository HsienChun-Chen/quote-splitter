"""最終版 Python 參考實作 - 用來驗證邏輯，之後要翻譯成 JS"""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from copy import copy
import subprocess, os

TEMPLATE = "/mnt/user-data/uploads/活頁簿1.xlsx"
FIRST_ROW = 4
MAX_ROWS = 15
LAST_ROW = FIRST_ROW + MAX_ROWS - 1

# D 欄可容納的大約字元數（用於計算列高）
DESC_COL_CHARS = 18  # 每行約 18 個中文字
BASE_ROW_HEIGHT = 18  # 一行高度


def split_pages(items, capacity):
    """切分多頁，換頁時帶出最後的分類列作為「續」"""
    pages = []
    current = []
    last_section = None
    total = len(items)
    for idx, item in enumerate(items):
        if item.get("is_section"):
            last_section = item
        current.append(item)
        # 只有在還有後續項目時才考慮切頁
        still_have_more = idx < total - 1
        if len(current) >= capacity and still_have_more:
            pages.append(current)
            current = []
            if last_section and not items[idx + 1].get("is_section"):
                current.append({"is_section": True, "desc": last_section["desc"] + "（續）"})
    if current:
        pages.append(current)
    return pages


def fill_sheet(ws, items, vendor_name, project, page_num, total_pages):
    ws["A2"] = ""  # 清開發者備註

    # === 強制頁面適合列印 ===
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = "B2:I47"

    # === 取得參考樣式（第 5 列是正常品項列範本）===
    ref_row = 5
    ref_styles = {}
    for col in ["B", "D", "E", "F", "G", "H"]:
        cell = ws[f"{col}{ref_row}"]
        ref_styles[col] = {
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "alignment": copy(cell.alignment),
            "border": copy(cell.border),
            "number_format": cell.number_format
        }

    # === 找出模板中原有的 B:I 合併列 ===
    template_section_rows = set()
    for mr in list(ws.merged_cells.ranges):
        s = str(mr)
        if s.startswith("B") and ":I" in s:
            row_num = int(s.split("B")[1].split(":")[0])
            if FIRST_ROW <= row_num <= LAST_ROW:
                template_section_rows.add(row_num)
                try:
                    ws.unmerge_cells(s)
                except Exception:
                    pass

    # === 把那些原分類列的樣式重設為品項列樣式（清除 medium 邊框殘留）===
    explicit_aligns = {
        "B": Alignment(horizontal="center", vertical="center"),
        "D": Alignment(vertical="center"),
        "E": Alignment(horizontal="right", vertical="center"),
        "F": Alignment(horizontal="right", vertical="center"),
        "G": Alignment(horizontal="right", vertical="center"),
        "H": Alignment(horizontal="right", vertical="center"),
    }

    for row in template_section_rows:
        for col in ["B", "D", "E", "F", "G", "H"]:
            cell = ws[f"{col}{row}"]
            s = ref_styles[col]
            cell.font = copy(s["font"])
            cell.fill = copy(s["fill"])
            cell.alignment = copy(explicit_aligns[col])
            cell.border = copy(s["border"])
            cell.number_format = s["number_format"]
        # C、I 欄（正常列是合併儲存格的第二格）清邊框
        for col in ["C", "I"]:
            cell = ws[f"{col}{row}"]
            cell.border = Border()
            cell.fill = PatternFill(fill_type=None)
        # 重新建立正常列的 B:C 和 H:I 合併（之後填品項時需要）
        # 這樣項次會置中顯示在 B:C 的中央
        try:
            ws.merge_cells(f"B{row}:C{row}")
            ws.merge_cells(f"H{row}:I{row}")
        except Exception:
            pass

    # === 分類列樣式 ===
    section_font = Font(name="Noto Sans TC DemiLight", size=12, bold=True)
    section_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    section_align = Alignment(horizontal="center", vertical="center")
    medium = Side(border_style="medium", color="000000")
    section_border = Border(top=medium, bottom=medium)

    # === 填資料 ===
    for i in range(MAX_ROWS):
        row = FIRST_ROW + i
        if i < len(items):
            item = items[i]
            if item.get("is_section"):
                # 要變成分類列：先 unmerge 該列所有合併，再合併 B:I
                for mr in list(ws.merged_cells.ranges):
                    s = str(mr)
                    if f"{row}" in s.split(":")[0] and s.split(":")[0][-len(str(row)):] == str(row):
                        # 該列起始的合併
                        import re
                        m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", s)
                        if m and int(m.group(2)) == row:
                            try: ws.unmerge_cells(s)
                            except: pass
                try:
                    ws.merge_cells(f"B{row}:I{row}")
                except Exception:
                    pass
                cell = ws[f"B{row}"]
                cell.value = item["desc"]
                cell.font = section_font
                cell.fill = section_fill
                cell.alignment = section_align
                cell.border = section_border
                # 分類列列高固定
                ws.row_dimensions[row].height = 20
            else:
                # 正常品項列，確保 B:C 和 H:I 合併
                try:
                    for mr in list(ws.merged_cells.ranges):
                        s = str(mr)
                        if s == f"B{row}:I{row}":
                            ws.unmerge_cells(s)
                            break
                    has_bc = any(str(mr) == f"B{row}:C{row}" for mr in ws.merged_cells.ranges)
                    if not has_bc:
                        ws.merge_cells(f"B{row}:C{row}")
                    has_hi = any(str(mr) == f"H{row}:I{row}" for mr in ws.merged_cells.ranges)
                    if not has_hi:
                        ws.merge_cells(f"H{row}:I{row}")
                except Exception:
                    pass

                ws[f"B{row}"] = item["no"]
                ws[f"D{row}"] = item["desc"]
                ws[f"E{row}"] = f"{float(item['qty']):.2f} {item['unit']}"
                ws[f"F{row}"] = "-"
                ws[f"G{row}"] = "-"
                ws[f"H{row}"] = item.get("note", "-")
                # 強制對齊
                for col in ["B", "D", "E", "F", "G", "H"]:
                    cell = ws[f"{col}{row}"]
                    cell.alignment = copy(explicit_aligns[col])
                # D 欄（品項）自動換行 + center 對齊
                ws[f"D{row}"].alignment = Alignment(vertical="center", wrap_text=True)
                for col in ["B", "E", "F", "G", "H"]:
                    cell = ws[f"{col}{row}"]
                    cur = cell.alignment
                    cell.alignment = Alignment(horizontal=cur.horizontal, vertical="center")
                # 根據描述文字長度估算所需列高
                desc_len = len(item["desc"])
                # 中文字約 18 個為一行；計算行數
                lines_needed = max(1, (desc_len + DESC_COL_CHARS - 1) // DESC_COL_CHARS)
                ws.row_dimensions[row].height = BASE_ROW_HEIGHT * lines_needed
        else:
            # 超出範圍的列：清空內容、清除邊框/填充
            for col in ["B", "D", "E", "F", "G", "H"]:
                ws[f"{col}{row}"] = None
            for col in ["B", "C", "D", "E", "F", "G", "H", "I"]:
                ws[f"{col}{row}"].border = Border()
                ws[f"{col}{row}"].fill = PatternFill(fill_type=None)
            # 把未使用的列高設為很小，讓版面緊湊
            ws.row_dimensions[row].height = 3

        for col in ["J", "K", "L"]:
            ws[f"{col}{row}"] = None

    # === 處理 19-28 列（範本中原本的品項列但超出 15 項範圍）===
    # 只有在 MAX_ROWS < 25 的情況下才需要處理
    for row in range(FIRST_ROW + MAX_ROWS, 29):  # 19-28
        # 取消 B:C 和 H:I 合併（如果有）
        for range_str in [f"B{row}:C{row}", f"H{row}:I{row}", f"B{row}:I{row}"]:
            try:
                ws.unmerge_cells(range_str)
            except Exception:
                pass
        # 清空內容和邊框
        for col in ["B", "D", "E", "F", "G", "H"]:
            try:
                ws[f"{col}{row}"] = None
            except Exception:
                pass
        for col in ["B", "C", "D", "E", "F", "G", "H", "I"]:
            try:
                ws[f"{col}{row}"].border = Border()
                ws[f"{col}{row}"].fill = PatternFill(fill_type=None)
            except Exception:
                pass
        # 列高壓縮到最小
        ws.row_dimensions[row].height = 3

    # === 專案資訊 ===
    ws["D37"] = f" 聯絡專員 | {project.get('contact','')}"
    ws["D38"] = f" 聯絡電話 | {project.get('tel','')}"
    ws["D39"] = f" 電子信箱 | {project.get('mail','')}"
    ws["D40"] = f" 工程名稱 | {project.get('name','')}"
    ws["D41"] = f" 工程地點 | {project.get('site','')}"

    # === 廠商名稱 + 日期 ===
    # LOGO 圖片橫跨 B2:H2，所以 I2 是唯一不被擋的位置
    # 廠商名稱放一列、日期放另一列（但只有 I2 一格可用，所以合併成一行用 | 分隔）
    from datetime import date
    today_str = date.today().strftime("%Y/%m/%d")
    vendor_text = vendor_name
    if total_pages > 1:
        vendor_text = f"{vendor_name} ({page_num}/{total_pages})"

    try:
        # 覆蓋 I2 的 TODAY() 公式
        ws["I2"] = f"{vendor_text}\n{today_str}"
        ws["I2"].font = Font(name="Noto Sans TC DemiLight", size=9, bold=True)
        ws["I2"].alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
    except Exception as e:
        print(f"warning setting vendor name: {e}")

    # 清其他列 J/K/L 公式 + B47 的外部連結公式
    for r in range(1, 50):
        for col in ["J", "K", "L"]:
            try:
                cell = ws[f"{col}{r}"]
                if cell.value and str(cell.value).startswith("="):
                    cell.value = None
            except Exception:
                pass

    # 清除所有儲存格中的外部連結公式（避免 #REF 或殘留的 "項次" 字樣）
    for row_cells in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=12):
        for cell in row_cells:
            try:
                v = cell.value
                if isinstance(v, str) and v.startswith("=") and "[1]" in v:
                    cell.value = None
            except Exception:
                pass

    # 第 47 列是範本的「合計」列（B:I 合併）
    try:
        ws["B47"] = ""
    except Exception:
        pass

    # 清除列 29-32 (原合計區) 的淡灰色背景
    for r in [29, 30, 31, 32]:
        for col in ["B", "C", "D", "E", "F", "G", "H", "I"]:
            try:
                ws[f"{col}{r}"].fill = PatternFill(fill_type=None)
            except Exception:
                pass


def build_quote(template_path, output_path, vendor_name, project, items):
    wb = load_workbook(template_path)
    original_sheet = wb.active
    pages = split_pages(items, MAX_ROWS)

    if len(pages) == 1:
        fill_sheet(original_sheet, pages[0], vendor_name, project, 1, 1)
        original_sheet.title = vendor_name[:30]
    else:
        fill_sheet(original_sheet, pages[0], vendor_name, project, 1, len(pages))
        original_sheet.title = f"{vendor_name}_P1"[:30]
        for p in range(1, len(pages)):
            copied = wb.copy_worksheet(original_sheet)
            copied.title = f"{vendor_name}_P{p+1}"[:30]
            fill_sheet(copied, pages[p], vendor_name, project, p + 1, len(pages))

    wb.save(output_path)


if __name__ == "__main__":
    proj = {
        "contact": "陳憲君",
        "tel": "0912-687-638",
        "mail": "a7788995720@gmail.com",
        "name": "屏東區管理處暨營運所辦公廳興建工程",
        "site": "屏東市林森路"
    }

    # Test 1: PVC 地磚（類似範例）
    items_1 = [
        {"is_section": True, "desc": "PVC地磚工程"},
        {"no": 1, "desc": "地坪,自平水泥+15*90cmPVC地磚A", "unit": "M2", "qty": 1402.00, "note": "-"},
        {"no": 2, "desc": "地坪,自平水泥+15*90cmPVC地磚B", "unit": "M2", "qty": 527.00, "note": "-"},
        {"no": 3, "desc": "側封板(H=20CM,面板橫向鋪設)", "unit": "M", "qty": 46.00, "note": "-"},
    ]
    build_quote(TEMPLATE, "/home/claude/final_1.xlsx", "A供應商", proj, items_1)

    # Test 2: 14 項（正好在 15 項容量內）
    items_2 = [{"is_section": True, "desc": "假設工程"}]
    for i in range(1, 15):
        items_2.append({"no": i, "desc": f"測試項目 {i}", "unit": "式", "qty": i * 10.5, "note": "-"})
    build_quote(TEMPLATE, "/home/claude/final_2.xlsx", "B供應商", proj, items_2)

    # Test 3: 25 項（要分兩頁 15+10）+ 長文字項目
    items_3 = [{"is_section": True, "desc": "甲、電氣工程"}]
    items_3.append({"no": 1, "desc": "新設單開橫拉式60A電動防火門 W150*H240 含自動感應開關及防夾裝置(電源由甲方另外提供)", "unit": "樘", "qty": 1, "note": "-"})
    items_3.append({"no": 2, "desc": "PVC導管及壓修", "unit": "式", "qty": 1, "note": "-"})
    items_3.append({"no": 3, "desc": "插座迴路(PVC電線2.0mm²)", "unit": "迴", "qty": 7, "note": "-"})
    items_3.append({"no": 4, "desc": "插座箱體另計", "unit": "式", "qty": 1, "note": "-"})
    items_3.append({"no": 5, "desc": "插座及其配管配線含各項相關設備連接及測試工作完成", "unit": "式", "qty": 1, "note": "-"})
    for i in range(6, 20):
        items_3.append({"no": i, "desc": f"測試項目 {i} 內容範例描述文字可能會有點長度", "unit": "式", "qty": i * 2.5, "note": "-"})
    items_3.append({"is_section": True, "desc": "乙、資訊工程"})
    for i in range(20, 26):
        items_3.append({"no": i, "desc": f"資訊項目 {i}", "unit": "組", "qty": 1, "note": "-"})
    build_quote(TEMPLATE, "/home/claude/final_3.xlsx", "C水電行", proj, items_3)

    # 全部轉 PDF 檢視
    for i in [1, 2, 3]:
        subprocess.run([
            "python3", "/mnt/skills/public/xlsx/scripts/office/soffice.py",
            "--convert-to", "pdf", "--outdir", "/home/claude",
            f"/home/claude/final_{i}.xlsx"
        ], capture_output=True)
        import pypdfium2 as pdfium
        pdf = pdfium.PdfDocument(f"/home/claude/final_{i}.pdf")
        print(f"final_{i}.pdf has {len(pdf)} pages")
        for j, page in enumerate(pdf):
            img = page.render(scale=1.4).to_pil()
            img.save(f"/home/claude/final_{i}_p{j+1}.png")
